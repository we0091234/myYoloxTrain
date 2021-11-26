#!/usr/bin/env python3
# -*- coding:utf-8 -*-
# Copyright (c) Megvii, Inc. and its affiliates.

import argparse
import os
import time
from loguru import logger

import cv2
import openpyxl
from openpyxl import workbook
import torch

from yolox.data.data_augment import ValTransform
from yolox.data.datasets import COCO_CLASSES
from yolox.data.datasets import VOC_CLASSES
from yolox.exp import get_exp
from yolox.utils import fuse_model, get_model_info, postprocess, vis

IMAGE_EXT = [".jpg", ".jpeg", ".webp", ".bmp", ".png"]

def allFilePath(rootPath,allFIleList):
    fileList = os.listdir(rootPath)
    for temp in fileList:
        if os.path.isfile(os.path.join(rootPath,temp)):
            if temp.endswith(".jpg"):
                allFIleList.append(os.path.join(rootPath,temp))
        else:
            allFilePath(os.path.join(rootPath,temp),allFIleList)

def allFilePathmodel(rootPath,allFIleList):
    fileList = os.listdir(rootPath)
    for temp in fileList:
        if os.path.isfile(os.path.join(rootPath,temp)):
            if temp.endswith(".pth"):
                allFIleList.append(os.path.join(rootPath,temp))
        else:
            allFilePath(os.path.join(rootPath,temp),allFIleList)

def make_parser():
    parser = argparse.ArgumentParser("YOLOX Demo!")
    parser.add_argument(
        "demo", default="image", help="demo type, eg. image, video and webcam"
    )
    parser.add_argument("-expn", "--experiment-name", type=str, default=None)
    parser.add_argument("-n", "--name", type=str, default=None, help="model name")

    parser.add_argument(
        "--model_path", default="./assets/dog.jpg", help="path to images or video"
    )
    parser.add_argument(
        "--image_path", default="./assets/dog.jpg", help="path to images or video"
    )
    parser.add_argument(
        "--excel_path", default="./assets/dog.jpg", help="path to images or video"
    )
    parser.add_argument("--camid", type=int, default=0, help="webcam demo camera id")
    parser.add_argument(
        "--save_result",
        action="store_true",
        help="whether to save the inference result of image/video",
    )

    # exp file
    parser.add_argument(
        "-f",
        "--exp_file",
        default=None,
        type=str,
        help="pls input your expriment description file",
    )
    parser.add_argument("-c", "--ckpt", default=None, type=str, help="ckpt for eval")
    parser.add_argument(
        "--device",
        default="cpu",
        type=str,
        help="device to run our model, can either be cpu or gpu",
    )
    parser.add_argument("--conf", default=0.3, type=float, help="test conf")
    parser.add_argument("--nms", default=0.3, type=float, help="test nms threshold")
    parser.add_argument("--tsize_h", default=None, type=int, help="test img size")
    parser.add_argument("--tsize_w", default=None, type=int, help="test img size")
    parser.add_argument(
        "--fp16",
        dest="fp16",
        default=False,
        action="store_true",
        help="Adopting mix precision evaluating.",
    )
    parser.add_argument(
        "--legacy",
        dest="legacy",
        default=False,
        action="store_true",
        help="To be compatible with older versions",
    )
    parser.add_argument(
        "--fuse",
        dest="fuse",
        default=False,
        action="store_true",
        help="Fuse conv and bn for testing.",
    )
    parser.add_argument(
        "--trt",
        dest="trt",
        default=False,
        action="store_true",
        help="Using TensorRT model for testing.",
    )
    return parser


def get_image_list(path):
    image_names = []
    for maindir, subdir, file_name_list in os.walk(path):
        for filename in file_name_list:
            apath = os.path.join(maindir, filename)
            ext = os.path.splitext(apath)[1]
            if ext in IMAGE_EXT:
                image_names.append(apath)
    return image_names


class Predictor(object):
    def __init__(
        self,
        model,
        exp,
        cls_names=COCO_CLASSES,
        trt_file=None,
        decoder=None,
        device="cpu",
        legacy=False,
    ):
        self.model = model
        self.cls_names = cls_names
        self.decoder = decoder
        self.num_classes = exp.num_classes
        self.confthre = exp.test_conf
        self.nmsthre = exp.nmsthre
        self.test_size = exp.test_size
        self.device = device
        self.preproc = ValTransform(legacy=legacy)
        if trt_file is not None:
            from torch2trt import TRTModule

            model_trt = TRTModule()
            model_trt.load_state_dict(torch.load(trt_file))

            x = torch.ones(1, 3, exp.test_size[0], exp.test_size[1]).cuda()
            self.model(x)
            self.model = model_trt

    def inference(self, img):
        # global sumtime
        img_info = {"id": 0}
        if isinstance(img, str):
            img_info["file_name"] = os.path.basename(img)
            img = cv2.imread(img)
            h,w,c=img.shape
        else:
            img_info["file_name"] = None

        height, width = img.shape[:2]
        img_info["height"] = height
        img_info["width"] = width
        img_info["raw_img"] = img

        ratio = min(self.test_size[0] / img.shape[0], self.test_size[1] / img.shape[1])
        img_info["ratio"] = ratio

        img, _ = self.preproc(img, None, self.test_size)
        img = torch.from_numpy(img).unsqueeze(0)
        if self.device == "gpu":
            img = img.cuda()

        with torch.no_grad():
            t0 = time.time()
            outputs = self.model(img)
            t1=time.time()
            # sumtime+=t1-t0
            logger.info("Infer time: {:.4f}s".format(t1 - t0))
            if self.decoder is not None:
                outputs = self.decoder(outputs, dtype=outputs.type())
            outputs = postprocess(
                outputs, self.num_classes, self.confthre,
                self.nmsthre, class_agnostic=True
            )
            
        return outputs, img_info,h,w,ratio, t1-t0

    def visual(self, output, img_info, cls_conf=0.35):
        ratio = img_info["ratio"]
        img = img_info["raw_img"]
        if output is None:
            return img
        output = output.cpu()

        bboxes = output[:, 0:4]

        # preprocessing: resize
        bboxes /= ratio

        cls = output[:, 6]
        scores = output[:, 4] * output[:, 5]

        vis_res = vis(img, bboxes, scores, cls, cls_conf, self.cls_names)
        return vis_res


def image_demo(predictor, vis_folder, path, current_time, save_result):
    if os.path.isdir(path):
        files = get_image_list(path)
    else:
        files = [path]
    files.sort()
    for image_name in files:
        outputs, img_info = predictor.inference(image_name)
        result_image = predictor.visual(outputs[0], img_info, predictor.confthre)
        if save_result:
            save_folder = os.path.join(
                vis_folder, time.strftime("%Y_%m_%d_%H_%M_%S", current_time)
            )
            os.makedirs(save_folder, exist_ok=True)
            save_file_name = os.path.join(save_folder, os.path.basename(image_name))
            logger.info("Saving detection result in {}".format(save_file_name))
            cv2.imwrite(save_file_name, result_image)
        ch = cv2.waitKey(0)
        if ch == 27 or ch == ord("q") or ch == ord("Q"):
            break


def imageflow_demo(predictor, vis_folder, current_time, args):
    cap = cv2.VideoCapture(args.path if args.demo == "video" else args.camid)
    width = cap.get(cv2.CAP_PROP_FRAME_WIDTH)  # float
    height = cap.get(cv2.CAP_PROP_FRAME_HEIGHT)  # float
    fps = cap.get(cv2.CAP_PROP_FPS)
    save_folder = os.path.join(
        vis_folder, time.strftime("%Y_%m_%d_%H_%M_%S", current_time)
    )
    os.makedirs(save_folder, exist_ok=True)
    if args.demo == "video":
        save_path = os.path.join(save_folder, args.path.split("/")[-1])
    else:
        save_path = os.path.join(save_folder, "camera.mp4")
    logger.info(f"video save_path is {save_path}")
    vid_writer = cv2.VideoWriter(
        save_path, cv2.VideoWriter_fourcc(*"mp4v"), fps, (int(width), int(height))
    )
    while True:
        ret_val, frame = cap.read()
        if ret_val:
            outputs, img_info = predictor.inference(frame)
            result_frame = predictor.visual(outputs[0], img_info, predictor.confthre)
            if args.save_result:
                vid_writer.write(result_frame)
            ch = cv2.waitKey(1)
            if ch == 27 or ch == ord("q") or ch == ord("Q"):
                break
        else:
            break
def writeToExcel(excelPath,modelPath,numClass,allCat,mylabel=None,args=None):
    str_xlsx =excelPath
    p1 = os.path.exists(str_xlsx)
    if p1:
        excel_out = openpyxl.load_workbook(str_xlsx)
    else:
        excel_out = workbook.Workbook()
    sheetnames = excel_out.get_sheet_names()
    table = excel_out.get_sheet_by_name(sheetnames[0])
    table = excel_out.active

    nrows = table.max_row
    nrows = nrows + 1
    table.cell(nrows, 1).value = modelPath+"_"+str(args.tsize_h)+"_"+str(args.tsize_h)+"_conf:"+str(args.conf)
    nrows = nrows + 1
    table.cell(nrows, 1).value =args.image_path
    nrows = nrows + 1
    list_1 = ["class", "src_num", "det_num", "right_num", "recall", "pre"]
    for i in range(len(list_1)):
        table.cell(nrows, i + 1).value = list_1[i]
    nrows = nrows + 1

    for i in range(numClass):
        list_2 = [mylabel[i], allCat[i].sum, allCat[i].detect, allCat[i].right, allCat[i].detectRatio, allCat[i].accuracy]
        for j in range(len(list_2)):
            table.cell(nrows, j + 1).value = list_2[j]
        nrows = nrows + 1
    excel_out.save(str_xlsx)

class Category():
    def __init__(self):
           self.detect=0
           self.sum = 0
           self.right=0
           self.miss=0
           self.detectRatio=0
           self.accuracy=0

def bbOverlap1(src_box,det_box):
    src_x1=src_box[0]
    src_y1 = src_box[1]
    src_x2 = src_box[2]
    src_y2 = src_box[3]
    det_x1 = det_box[0]
    det_y1 = det_box[1]
    det_x2 = det_box[2]
    det_y2 = det_box[3]

    if src_x1 > det_x2 :
         return 0.0
    if src_y1 > det_y2:
        return 0.0
    if src_x2 < det_x1:
        return 0.0
    if src_y2 < det_y1 :
        return 0.0
    colInt = min(src_x2, det_x2) - max(src_x1, det_x1)
    rowInt = min(src_y2, det_y2) - max(src_y1, det_y1)

    intersection = colInt * rowInt
    area1 = (src_x2-src_x1) * (src_y2-src_y1)
    area2 = (det_x2-det_x1) * (det_y2-det_y1)
    return intersection / (area1 + area2 - intersection)

def getTxtLabel(textfileNmae,width,height):
    bboxs=[]
    if not os.path.exists(textfileNmae):
        return bboxs
    with open(textfileNmae) as f:
        while True:
            box = []
            line = f.readline()
            if not line:
                break
            lineList = line.split()
            if len(lineList) < 5:
                continue
            ww = float(lineList[3]) * width
            hh = float(lineList[4]) * height
            xx = float(lineList[1]) * width - ww / 2
            yy = float(lineList[2]) * height - hh / 2

            box.append(int(xx))
            box.append(int(yy))
            box.append(int(xx + ww))
            box.append(int(yy + hh))
            box.append(int(lineList[0]))
            # f1.write(name+" "+str(int(xx))+" "+str(int(yy))+" "+str(int(ww))+" "+str(int(hh))+"\n")
            # print(box)
            bboxs.append(box)
    f.close()
    return bboxs

def getBox(predictor,image_name):
    new_dets=[]
    outputs, _ ,h,w,ratio,inferTime= predictor.inference(image_name)
    if outputs[0] is None:
        return new_dets,h,w,inferTime
    output=outputs[0].cpu().numpy()
    for i in range(output.shape[0]):
        oriDet = output[i].tolist()
        deti=oriDet[0:4]
        deti=[x/ratio for x in deti]
        deti=[int(x) for x in deti]
        deti.append(int (oriDet[6]))
        new_dets.append(deti)

    return new_dets,h,w,inferTime

def main(exp, args,modelFile):
    global sumtime
    if not args.experiment_name:
        args.experiment_name = exp.exp_name

    file_name = os.path.join(exp.output_dir, args.experiment_name)
    os.makedirs(file_name, exist_ok=True)

    vis_folder = None
    if args.save_result:
        vis_folder = os.path.join(file_name, "vis_res")
        os.makedirs(vis_folder, exist_ok=True)

    if args.trt:
        args.device = "gpu"

    logger.info("Args: {}".format(args))

    if args.conf is not None:
        exp.test_conf = args.conf
    if args.nms is not None:
        exp.nmsthre = args.nms
    if args.tsize_h is not None:
        exp.test_size = (args.tsize_h, args.tsize_w)
        # exp.test_size = (160,96)
        

    model = exp.get_model()
    logger.info("Model Summary: {}".format(get_model_info(model, exp.test_size)))

    if args.device == "gpu":
        model.cuda()
    model.eval()

    if not args.trt:
        if args.ckpt is None:
            ckpt_file = os.path.join(file_name, "best_ckpt.pth")
        else:
            # ckpt_file = args.ckpt
            ckpt_file = modelFile
        logger.info("loading checkpoint")
        ckpt = torch.load(ckpt_file, map_location="cpu")
        # load the model state dict
        model.load_state_dict(ckpt["model"])
        logger.info("loaded checkpoint done.")

    if args.fuse:
        logger.info("\tFusing model...")
        model = fuse_model(model)

    if args.trt:
        assert not args.fuse, "TensorRT model is not support model fusing!"
        trt_file = os.path.join(file_name, "model_trt.pth")
        assert os.path.exists(
            trt_file
        ), "TensorRT model is not found!\n Run python3 tools/trt.py first!"
        model.head.decode_in_inference = False
        decoder = model.head.decode_outputs
        logger.info("Using TensorRT to inference")
    else:
        trt_file = None
        decoder = None

    predictor = Predictor(model, exp, COCO_CLASSES, trt_file, decoder, args.device, args.legacy)

    imagePath =args.image_path
    # imagePath =r"G:/detectData/two_car"
    excelPath= args.excel_path
    # modelPath =r"M:\Yolox\YOLOX-main\model_s"
    mylabel=['head', 'leg', 'hand','back','nostd',"body","plate","logo",]
    numClass=8
    fileList=[]
    allFilePath(imagePath,fileList)
    allCat = []
    for i in range(numClass):
        OneCat = Category()
        allCat.append(OneCat)
    picCount=0
    for image_name in fileList:
       picCount+=1
       print(picCount,image_name)
       detBox,h,w,inferTime=getBox(predictor,image_name)
       if (picCount!=1):
        sumtime+=inferTime
       txtFile=image_name.replace(".jpg",".txt")
       srcbboxs= getTxtLabel(txtFile,w,h)
    #    print(detBox,srcbboxs)
       
       for box in detBox:
            allCat[box[4]].detect+=1
        # print(bboxs)
       for srcbox in srcbboxs:
            allCat[srcbox[4]].sum+=1
            flag=0
            for detbox in detBox:
                if detbox[4]==srcbox[4] and bbOverlap1(srcbox,detbox)>0.5:
                    allCat[srcbox[4]].right+=1
                    flag=1
                    break
            if not flag:
                allCat[srcbox[4]].miss+=1

    for i in range(len(allCat)):
        if allCat[i].sum == 0 or allCat[i].detect==0:
            continue
        allCat[i].detectRatio = 1.0 * allCat[int(i)].right / allCat[int(i)].sum
        allCat[i].accuracy=1.0 * allCat[int(i)].right / allCat[int(i)].detect

    for i in range(len(allCat)):
        print("class:{} sum:{} right:{} miss:{} detectRatio:{} detect:{} accuracy:{}".format(mylabel[i],allCat[i].sum,allCat[i].right,allCat[i].miss,allCat[i].detectRatio,allCat[i].detect,allCat[i].accuracy))
    writeToExcel(excelPath,modelFile,numClass,allCat,mylabel,args)
    print("sumTime={} ms,average={} ms".format(sumtime*1000,(sumtime/len(fileList))*1000))
    # current_time = time.localtime()
    # current_time = time.localtime()
    # if args.demo == "image":
    #     image_demo(predictor, vis_folder, args.path, current_time, args.save_result)
    # elif args.demo == "video" or args.demo == "webcam":
    #     imageflow_demo(predictor, vis_folder, current_time, args)

sumtime = 0
if __name__ == "__main__":
    args = make_parser().parse_args()
    exp = get_exp(args.exp_file, args.name)
    # modelfolder = r"/mnt/Gpan/Mydata/pytorchPorject/yoloxNew/myYoloxTrain/modelTest1"
    modelfolder = args.model_path
    modelFileList =[]
    allFilePathmodel(modelfolder,modelFileList)
    for modelFile in modelFileList:
        main(exp, args,modelFile)
